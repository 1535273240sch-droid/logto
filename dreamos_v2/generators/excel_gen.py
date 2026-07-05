"""Excel 生成器 — 生成 .xlsx 文件"""
from pathlib import Path
from . import BaseGenerator


class ExcelGenerator(BaseGenerator):
    supported_types = ["excel", "csv"]

    async def generate(self, content: str, output_path: Path,
                       title: str = "", **kwargs) -> dict:
        self._ensure_dir(output_path)
        artifact_type = kwargs.get("artifact_type", "excel")

        if artifact_type == "csv":
            return await self._gen_csv(content, output_path, title)
        return await self._gen_excel(content, output_path, title)

    async def _gen_excel(self, content: str, output_path: Path,
                         title: str) -> dict:
        """生成 Excel 文件"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment

            wb = Workbook()
            ws = wb.active
            ws.title = title[:31] if title else "Sheet1"

            # 样式
            header_font = Font(bold=True, color="FFFFFF", size=12)
            header_fill = PatternFill(start_color="6c5ce7", end_color="6c5ce7",
                                      fill_type="solid")
            cell_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

            # 解析内容为表格
            rows = self._parse_content(content)

            # 写入数据
            for row_idx, row in enumerate(rows):
                for col_idx, cell_value in enumerate(row):
                    cell = ws.cell(row=row_idx + 1, column=col_idx + 1, value=cell_value)
                    cell.alignment = cell_align
                    if row_idx == 0:  # 表头
                        cell.font = header_font
                        cell.fill = header_fill

            # 自适应列宽
            for col in ws.columns:
                max_len = max(len(str(c.value or "")) for c in col)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

            wb.save(output_path)
            return {"success": True}

        except ImportError:
            # openpyxl 不可用时降级为 CSV
            return await self._gen_csv(content, output_path.with_suffix(".csv"), title)

    async def _gen_csv(self, content: str, output_path: Path,
                       title: str) -> dict:
        """生成 CSV 文件"""
        import csv
        rows = self._parse_content(content)

        with open(output_path, "w", encoding="utf-8-sig", newline="") as f:
            writer = csv.writer(f)
            for row in rows:
                writer.writerow(row)

        return {"success": True}

    def _parse_content(self, content: str) -> list[list[str]]:
        """解析内容为二维表格"""
        rows = []
        for line in content.strip().split("\n"):
            line = line.strip()
            if not line or line.startswith("#") or line.startswith("---"):
                continue

            # Markdown 表格格式
            if "|" in line:
                cells = [c.strip() for c in line.split("|")]
                cells = [c for c in cells if c and c != "---"]
                if cells:
                    rows.append(cells)
            # 列表格式
            elif line.startswith("-") or line.startswith("*") or line.startswith("•"):
                clean = line.lstrip("-*• ").strip()
                rows.append([clean])
            # 有序格式
            elif len(line) > 2 and line[0].isdigit() and "." in line[:4]:
                clean = line.split(".", 1)[1].strip()
                rows.append([clean])
            # "键: 值" 格式
            elif ":" in line or "：" in line:
                sep = ":" if ":" in line else "："
                parts = line.split(sep, 1)
                rows.append([parts[0].strip(), parts[1].strip()])
            else:
                rows.append([line])

        # 确保所有行的列数一致
        if rows:
            max_cols = max(len(r) for r in rows)
            for row in rows:
                while len(row) < max_cols:
                    row.append("")

        return rows
